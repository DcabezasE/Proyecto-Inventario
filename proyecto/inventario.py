import datetime
import openpyxl

class Producto:
    def __init__(self, nombre: str, precio: float, cantidad: int, tiempo: str):
        self.nombre = nombre
        self.precio = precio
        self.cantidad = cantidad
        self.tiempo = tiempo
        
class Inventario:
    def __init__(self):
        self.archivo_excel = 'inventario.xlsx'
        self.productos = []
        self.fecha_hora= datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        self.guardar_datos()

    def registrar_entrada(self, producto, cantidad):
        for i in self.productos:
            if i.nombre == producto.nombre:
                i.cantidad += cantidad
                break
        else:
            self.productos.append(producto)
        print(f"Registro de entrada: {producto.nombre} x{cantidad} - {producto.tiempo}")

        self.guardar_datos()
        self.productos.clear()

    def registrar_salida(self, producto, cantidad):
        for i in self.productos:
            if i.nombre == producto.nombre:
                if i.cantidad >= cantidad:
                    i.cantidad -= cantidad
                    self.guardar_datos()
                    print(f"Registro de salida: {producto.nombre} x{cantidad} - {datetime.date.today()}")
                else:
                    print("No hay suficiente stock")
                break
        else:
            print("Producto no encontrado")

    def guardar_datos(self):
        try:
            workbook = openpyxl.load_workbook(self.archivo_excel)
            sheet = workbook['Inventario']
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Inventario'
            # Escribir encabezados si el archivo no existe
            sheet.append(['Nombre', 'Precio', 'Cantidad', 'Fecha y Hora'])

        # Crear un diccionario para buscar productos existentes por nombre
        productos_existentes = {row[0]: idx + 2 for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True))}
        print(productos_existentes)
        for producto in self.productos:
            if producto.nombre in productos_existentes:
                print("Ya existia")
                # Si el producto ya existe, actualizar su cantidad y tiempo

                fila = productos_existentes[producto.nombre]
                print(fila)
                nueva_cantidad = sheet[f"C{fila}"].value + producto.cantidad
                sheet[f"C{fila}"] = nueva_cantidad
                sheet[f"D{fila}"] = producto.tiempo
            else:
                print("No existia")
                # Si el producto es nuevo, agregarlo
                sheet.append([producto.nombre, producto.precio, producto.cantidad, producto.tiempo])

        workbook.save(self.archivo_excel)

    def cargar_datos(self):
        try:
            workbook = openpyxl.load_workbook(self.archivo_excel)
            sheet = workbook['Inventario']

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre, precio, cantidad, tiempo = row
                producto = Producto(nombre, precio, cantidad, tiempo)
                self.productos.append(producto)
        except FileNotFoundError:
            # Si el archivo no existe, no hacer nada
            pass
