#Empezamos en este modulo importando datetime para las fechas y hora de registro, openpyxl para manejar archivos xslx (excel) y el messagebox de tkinter (ya explico bien ahorita eso)
import datetime
import openpyxl
from tkinter import messagebox
#Se crea una clase Producto para instaciar los items de la empresa
class Producto:
    #La estructura del producto es: su nombre, su precio, cuanto hay en almacenado, y cuando se hizo el registro de ese producto
    def __init__(self, nombre: str, precio: float, cantidad: int, tiempo: str):
        self.nombre = nombre
        self.precio = precio
        self.cantidad = cantidad
        self.tiempo = tiempo

#Se crea la clase de Inventario que tiene relacion con la clase Producto en forma de composición
class Inventario:
    #Se estructura el nombre del archivo xlsx, la lista que almacenará los items, el formato del tiempo, y el metodo guardar_datos(ya explico eso)
    def __init__(self):
        self.archivo_excel = 'inventario.xlsx'
        self.productos = []
        self.fecha_hora= datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        self.guardar_datos()
    #Este metodo se encarga de ingresar los productos en la lista del inventario mediante el metodo cargar_datos, si existe en el inventario le suma la cantidad, sino lo crea, luego guarda y limpia
    def registrar_entrada(self, producto, cantidad):
        self.cargar_datos()
        for i in self.productos:
            if i.nombre == producto.nombre:
                i.cantidad += cantidad
                break
        else:
            self.productos.append(producto)
        print(f"Registro de entrada: {producto.nombre} x{cantidad} - {producto.tiempo}")

        self.guardar_datos()
        self.productos.clear()
    #Este metodo se encarga de la salida, carga los datos, y como si fuera una caja de market pasa la lisa_venta por un For para validar, disminuir la cantidad, guardar datos y llama a registrar_venta
    def registrar_salida(self, lista_venta):
        self.cargar_datos()
        for j in lista_venta:
            for i in self.productos:
                if i.nombre == j[0]:
                    if i.cantidad >= j[1]:
                        i.cantidad -= j[1]
                        self.guardar_datos()
                        self.registrar_venta(lista_venta)
                        print(f"Registro de salida: {j[0]} x{j[1]} - {datetime.date.today()}")
                    else:
                        print("No hay suficiente stock")
                    break
            else:
                print("Producto no encontrado")
    #Este metodo sirve para crear una hoja de ventas en el excel
    def registrar_venta(self, lista_venta):
        try:
            #Cargar el archivo Excel
            workbook = openpyxl.load_workbook(self.archivo_excel)
            #Verifica si la hoja "Ventas" ya existe, si no, crearla
            if 'Ventas' not in workbook.sheetnames:
                sheet = workbook.create_sheet('Ventas')
                #Crea encabezados para la hoja de ventas
                sheet.append(['Fecha y Hora', 'Producto', 'Cantidad', 'Precio Unitario', 'Total'])
            else:
                sheet = workbook['Ventas']
            #Recorre los productos vendidos y registrarlos en la hoja
            for item in lista_venta:
                producto, cantidad, precio_unitario = item
                total = cantidad * precio_unitario
                #Añade la fila de la venta
                sheet.append([datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S'), producto, cantidad, precio_unitario, total])
            #Guarda el archivo Excel
            workbook.save(self.archivo_excel)
            
        except Exception as e:
            #Muestra una notificacion de error y avisa porque fue
            messagebox.showerror("Error", f"Error al registrar la venta: {str(e)}")
    #El susodicho metodo de guardar_datos sirve para guardar los items recien añadidos
    def guardar_datos(self):
        try:
            #Carga el archivo y si no existe, lo crea
            workbook = openpyxl.load_workbook(self.archivo_excel)
            sheet = workbook['Inventario']
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Inventario'
            #Escribe encabezados si el archivo no existe
            sheet.append(['Nombre', 'Precio', 'Cantidad', 'Fecha y Hora'])

        #Crea un diccionario para buscar productos existentes por nombre e index
        productos_existentes = {row[0]: indx + 2 for indx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True))}
        #Se pasa producto por producto para saber cual es nuevo y cual actualizar el stock y eltiempo
        for producto in self.productos:
            if producto.nombre in productos_existentes:
                print("Ya existia")
                fila = productos_existentes[producto.nombre]
                nueva_cantidad = producto.cantidad
                sheet[f"C{fila}"] = nueva_cantidad
                sheet[f"D{fila}"] = producto.tiempo
            else:
                print("No existia")
                #Si el producto es nuevo, agregarlo
                sheet.append([producto.nombre, producto.precio, producto.cantidad, producto.tiempo])
        #Y guarda el archivo
        workbook.save(self.archivo_excel)
    #Este metodo carga el archivo para pasar todos la informacion a la lista de inventario y poder manipularla
    def cargar_datos(self):
        try:
            workbook = openpyxl.load_workbook(self.archivo_excel)
            sheet = workbook['Inventario']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre, precio, cantidad, tiempo = row
                producto = Producto(nombre, precio, cantidad, tiempo)
                self.productos.append(producto)
        except FileNotFoundError:
            #Si el archivo no existe, no hacer ni monda
            pass


    

